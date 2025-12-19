from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Optional

from fastapi import status
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session, joinedload

from app.core.errors import AppError
from app.models.export_log import ExportLog
from app.models.record import Record
from app.schemas.auth import SessionPayload
from app.services.external import ExternalDataAdapter
from app.services.utils import clean_value, first_value
from app.services.validation import ValidationService


def _date_range_to_window(from_date: date, to_date: date) -> tuple[datetime, datetime]:
    if to_date < from_date:
        raise AppError(code="validation_failed", message="时间范围不合法（to < from）", http_status=status.HTTP_422_UNPROCESSABLE_ENTITY)
    from_dt = datetime(from_date.year, from_date.month, from_date.day)
    to_dt = datetime(to_date.year, to_date.month, to_date.day) + timedelta(days=1)
    return from_dt, to_dt


def _fmt_date(value: Any) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def _quote_en(value: Any) -> Optional[str]:
    text = clean_value(value)
    if text is None:
        return None
    s = str(text)
    if s.startswith('"') and s.endswith('"') and len(s) >= 2:
        return s
    s = s.replace('"', '""')
    return f"\"{s}\""


def _unique_preserve_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for v in values:
        if v in seen:
            continue
        seen.add(v)
        result.append(v)
    return result


def _get_patient_no(row: dict[str, Any]) -> Optional[str]:
    value = first_value(row, ["JZKH", "jzkh", "BLH", "blh", "PATIENT_NO", "patient_no"])
    return clean_value(value)


def _fee_attr_for_code(code: str) -> str:
    # 兼容历史字段命名：BDBLZPF/QDBLZPF 在业务侧字段名为 bdbblzpf/qdbblzpf
    if code == "ZYL_ZYZD":
        return "zyl_zyzd"
    if code == "BDBLZPF":
        return "bdbblzpf"
    if code == "QDBLZPF":
        return "qdbblzpf"
    return code.lower()


@dataclass(frozen=True)
class ExportResult:
    filename: str
    content: bytes


class ExportService:
    def __init__(self, db: Session, external: ExternalDataAdapter) -> None:
        self.db = db
        self.external = external
        project_root = Path(__file__).resolve().parents[3]
        self._template_path = project_root / "3、中医门（急）诊诊疗信息页数据上传模板.xlsx"
        self._template_sheet = "门（急）诊诊疗数据上传表头"

    def export_report(self, *, from_date: date, to_date: date, session: SessionPayload) -> ExportResult:
        self._ensure_export_role(session)

        operator = session.doc_code or session.login_name
        from_dt, to_dt = _date_range_to_window(from_date, to_date)
        file_name = f"mz_mfp_report_{from_date.isoformat()}_{to_date.isoformat()}.xlsx"

        try:
            expected_patient_nos = self._fetch_patient_nos(from_dt=from_dt, to_dt=to_dt)
            if not expected_patient_nos:
                content = self._build_xlsx_bytes([], file_name=file_name)
                self._log_export(
                    status_="success",
                    created_by=operator,
                    file_name=file_name,
                    detail={"from": from_date.isoformat(), "to": to_date.isoformat(), "rows": 0},
                )
                self.db.commit()
                return ExportResult(filename=file_name, content=content)

            records = self._load_records(expected_patient_nos)
            by_patient = {r.patient_no: r for r in records}
            missing = [pn for pn in expected_patient_nos if pn not in by_patient]
            not_submitted = [pn for pn in expected_patient_nos if pn in by_patient and by_patient[pn].status != "submitted"]
            if missing or not_submitted:
                self._log_export(
                    status_="failed",
                    created_by=operator,
                    file_name=file_name,
                    error_message="存在未提交或未生成记录，阻断导出",
                    detail={"from": from_date.isoformat(), "to": to_date.isoformat(), "missing": len(missing), "not_submitted": len(not_submitted)},
                )
                self.db.commit()
                raise AppError(
                    code="validation_failed",
                    message="导出阻断：范围内存在未提交/缺失记录",
                    http_status=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail={"missing": missing, "not_submitted": not_submitted},
                )

            # 再次执行全量校验（规则可能迭代）
            validator = ValidationService(self.db)
            bad: list[dict[str, Any]] = []
            for pn in expected_patient_nos:
                rec = by_patient[pn]
                errs = validator.validate_for_submit(rec)
                if errs:
                    bad.append(
                        {
                            "patient_no": pn,
                            "record_id": rec.id,
                            "errors": [e.model_dump() for e in errs],
                        }
                    )
            if bad:
                self._log_export(
                    status_="failed",
                    created_by=operator,
                    file_name=file_name,
                    error_message="校验失败，阻断导出",
                    detail={"from": from_date.isoformat(), "to": to_date.isoformat(), "bad_records": len(bad)},
                )
                self.db.commit()
                raise AppError(
                    code="validation_failed",
                    message="导出阻断：存在校验错误",
                    http_status=status.HTTP_422_UNPROCESSABLE_ENTITY,
                    detail={"errors": bad},
                )

            ordered_records = [by_patient[pn] for pn in expected_patient_nos]
            content = self._build_xlsx_bytes(ordered_records, file_name=file_name)
            self._log_export(
                status_="success",
                created_by=operator,
                file_name=file_name,
                detail={"from": from_date.isoformat(), "to": to_date.isoformat(), "rows": len(ordered_records)},
            )
            self.db.commit()
            return ExportResult(filename=file_name, content=content)
        except AppError:
            raise
        except Exception as exc:
            self._log_export(
                status_="failed",
                created_by=operator,
                file_name=file_name,
                error_message="导出异常",
                detail={"type": type(exc).__name__},
            )
            self.db.commit()
            raise AppError(code="internal_error", message="导出失败，请稍后重试", http_status=status.HTTP_500_INTERNAL_SERVER_ERROR) from exc

    def _ensure_export_role(self, session: SessionPayload) -> None:
        if not any(role in {"admin", "qc"} for role in session.roles):
            raise AppError(code="forbidden", message="无权执行批量导出", http_status=status.HTTP_403_FORBIDDEN)

    def _fetch_patient_nos(self, *, from_dt: datetime, to_dt: datetime) -> list[str]:
        rows = self.external.fetch_visit_list(from_dt=from_dt, to_dt=to_dt)
        patient_nos: list[str] = []
        for raw in rows:
            pn = _get_patient_no(raw)
            if pn:
                patient_nos.append(pn)
        return _unique_preserve_order(patient_nos)

    def _load_records(self, patient_nos: list[str]) -> list[Record]:
        if not patient_nos:
            return []
        stmt = (
            select(Record)
            .where(Record.patient_no.in_(patient_nos))
            .options(
                joinedload(Record.org),
                joinedload(Record.base_info),
                joinedload(Record.diagnoses),
                joinedload(Record.tcm_operations),
                joinedload(Record.surgeries),
                joinedload(Record.medication_summary),
                joinedload(Record.herb_details),
                joinedload(Record.fee_summary),
            )
        )
        return list(self.db.execute(stmt).unique().scalars().all())

    def _load_template_headers(self) -> list[str]:
        if not self._template_path.exists():
            raise AppError(code="internal_error", message="导出模板文件不存在", http_status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        wb = load_workbook(self._template_path, read_only=True, data_only=True)
        if self._template_sheet in wb.sheetnames:
            ws = wb[self._template_sheet]
        else:
            ws = wb.worksheets[0]
        headers = [str(v).strip() for v in list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0] if v is not None]
        if not headers:
            raise AppError(code="internal_error", message="导出模板表头为空", http_status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        return headers

    def _build_xlsx_bytes(self, records: list[Record], *, file_name: str) -> bytes:
        headers = self._load_template_headers()
        wb = Workbook()
        ws = wb.active
        ws.title = self._template_sheet

        ws.append(headers)
        for record in records:
            row_map = self._record_to_row(record, headers)
            ws.append([row_map.get(h) for h in headers])

        tmp_path = None
        try:
            tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
            tmp_path = tmp.name
            tmp.close()
            wb.save(tmp_path)
            return Path(tmp_path).read_bytes()
        finally:
            if tmp_path:
                try:
                    Path(tmp_path).unlink(missing_ok=True)
                except Exception:
                    pass

    def _record_to_row(self, record: Record, headers: list[str]) -> dict[str, Any]:
        base = record.base_info
        org = record.org
        if base is None or org is None:
            raise AppError(code="internal_error", message="记录数据不完整", http_status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        data: dict[str, Any] = {h: None for h in headers}

        # 基础/就诊过程
        data["JGMC"] = clean_value(org.jgmc)
        data["ZZJGDM"] = clean_value(org.zzjgdm)
        data["USERNAME"] = clean_value(base.username)
        data["JZKH"] = clean_value(base.jzkh)
        data["XM"] = clean_value(base.xm)
        data["JZSJ"] = _fmt_date(base.jzsj)
        data["XB"] = clean_value(base.xb)
        data["CSRQ"] = _fmt_date(base.csrq)
        data["HY"] = clean_value(base.hy)
        data["GJ"] = clean_value(base.gj)
        data["MZ"] = clean_value(base.mz)
        data["ZJLB"] = clean_value(base.zjlb)
        data["ZJHM"] = clean_value(base.zjhm)
        data["XZZ"] = clean_value(base.xzz)
        data["LXDH"] = clean_value(base.lxdh)
        data["YWGMS"] = clean_value(base.ywgms)
        data["GMYW"] = _quote_en(base.gmyw)
        data["QTGMS"] = clean_value(base.qtgms)
        data["QTGMY"] = _quote_en(base.qtgmy)
        data["GHSJ"] = _fmt_date(base.ghsj)
        data["BDSJ"] = _fmt_date(base.bdsj)
        data["JZKS"] = clean_value(base.jzks)
        data["JZKSDM"] = clean_value(base.jzksdm)
        data["JZYS"] = clean_value(base.jzys)
        data["JZYSZC"] = clean_value(base.jzyszc)
        data["JZLX"] = clean_value(base.jzlx)
        data["FZ"] = clean_value(base.fz)
        data["SY"] = clean_value(base.sy)
        data["MZMTBHZ"] = clean_value(base.mzmtbhz)
        data["JZHZFJ"] = clean_value(base.jzhzfj)
        data["JZHZQX"] = clean_value(base.jzhzqx)
        data["ZYZKJSJ"] = _fmt_date(base.zyzkjsj)
        data["HZZS"] = _quote_en(base.hzzs)

        # 诊断映射
        diags = record.diagnoses or []
        def _diag(diag_type: str) -> list[Any]:
            return sorted([d for d in diags if d.diag_type == diag_type], key=lambda d: int(d.seq_no))

        tcm_dis = _diag("tcm_disease_main")
        if tcm_dis:
            data["MZD_ZB"] = clean_value(tcm_dis[0].diag_name)
            data["JBDM_ZB"] = clean_value(tcm_dis[0].diag_code)

        tcm_syn = _diag("tcm_syndrome")
        for idx in range(1, 3):
            item = tcm_syn[idx - 1] if idx - 1 < len(tcm_syn) else None
            data[f"MZD_ZZ{idx}"] = clean_value(item.diag_name) if item else None
            data[f"JBDM_ZZ{idx}"] = clean_value(item.diag_code) if item else None

        wm_main = _diag("wm_main")
        if wm_main:
            data["MZZD_ZYZD"] = clean_value(wm_main[0].diag_name)
            data["MZZD_JBBM"] = clean_value(wm_main[0].diag_code)

        wm_other = _diag("wm_other")
        for idx in range(1, 11):
            item = wm_other[idx - 1] if idx - 1 < len(wm_other) else None
            data[f"MZZD_QTZD{idx}"] = clean_value(item.diag_name) if item else None
            data[f"MZZD_JBBM{idx}"] = clean_value(item.diag_code) if item else None

        # 中医治疗性操作（≤10）
        ops = sorted(record.tcm_operations or [], key=lambda o: int(o.seq_no))
        for idx in range(1, 11):
            op = ops[idx - 1] if idx - 1 < len(ops) else None
            data[f"ZYZLCZMC{idx}"] = clean_value(op.op_name) if op else None
            data[f"ZYZLCZBM{idx}"] = clean_value(op.op_code) if op else None
            data[f"ZYZLCZCS{idx}"] = int(op.op_times) if op and op.op_times is not None else None
            data[f"ZYZLCZTS{idx}"] = int(op.op_days) if op and op.op_days is not None else None

        # 手术/操作（≤5）
        surgeries = sorted(record.surgeries or [], key=lambda s: int(s.seq_no))
        for idx in range(1, 6):
            s = surgeries[idx - 1] if idx - 1 < len(surgeries) else None
            data[f"SSCZMC{idx}"] = clean_value(s.op_name) if s else None
            data[f"SSCZBM{idx}"] = clean_value(s.op_code) if s else None
            data[f"SSCZRQ{idx}"] = _fmt_date(s.op_time) if s else None
            data[f"SSCZZ{idx}"] = clean_value(s.operator_name) if s else None
            data[f"MZFS{idx}"] = clean_value(s.anesthesia_method) if s else None
            data[f"MZYS{idx}"] = clean_value(s.anesthesia_doctor) if s else None
            data[f"SHJB{idx}"] = int(s.surgery_level) if s else None

        # 用药标识
        med = record.medication_summary
        data["XYSY"] = clean_value(getattr(med, "xysy", None))
        data["ZCYSY"] = clean_value(getattr(med, "zcysy", None))
        data["ZYZJSY"] = clean_value(getattr(med, "zyzjsy", None))
        data["CTYPSY"] = clean_value(getattr(med, "ctypsy", None))
        data["PFKLSY"] = clean_value(getattr(med, "pfklsy", None))

        herbs = sorted(record.herb_details or [], key=lambda h: int(h.seq_no))
        for idx in range(1, 41):
            h = herbs[idx - 1] if idx - 1 < len(herbs) else None
            data[f"ZCYLB{idx}"] = clean_value(h.herb_type) if h else None
            data[f"YYTJDM{idx}"] = clean_value(h.route_code) if h else None
            data[f"YYTJMC{idx}"] = clean_value(h.route_name) if h else None
            data[f"YYJS{idx}"] = int(h.dose_count) if h else None

        # 费用
        fee = record.fee_summary
        if fee is not None:
            for code in [
                "ZFY",
                "ZFJE",
                "YLFWF",
                "ZLCZF",
                "HLF",
                "QTFY",
                "BLZDF",
                "ZDF",
                "YXXZDF",
                "LCZDXMF",
                "FSSZLXMF",
                "ZLF",
                "SSZLF",
                "MZF",
                "SSF",
                "KFF",
                "ZYL_ZYZD",
                "ZYZL",
                "ZYWZ",
                "ZYGS",
                "ZCYJF",
                "ZYTNZL",
                "ZYGCZL",
                "ZYTSZL",
                "ZYQT",
                "ZYTSTPJG",
                "BZSS",
                "XYF",
                "KJYWF",
                "ZCYF",
                "ZYZJF",
                "ZCYF1",
                "PFKLF",
                "XF",
                "BDBLZPF",
                "QDBLZPF",
                "NXYZLZPF",
                "XBYZLZPF",
                "JCYYCLF",
                "YYCLF",
                "SSYCXCLF",
                "QTF",
            ]:
                attr = _fee_attr_for_code(code)
                value = getattr(fee, attr, None)
                if value is None:
                    data[code] = None
                elif isinstance(value, Decimal):
                    data[code] = float(value)
                else:
                    data[code] = value

        return data

    def _log_export(
        self,
        *,
        status_: str,
        created_by: str,
        file_name: str,
        error_message: Optional[str] = None,
        detail: Optional[dict[str, Any]] = None,
    ) -> None:
        self.db.add(
            ExportLog(
                record_id=None,
                export_type="xlsx",
                file_name=file_name,
                file_path=None,
                status=status_,
                error_message=error_message,
                detail=detail,
                created_by=created_by,
            )
        )
