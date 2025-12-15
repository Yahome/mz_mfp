from openpyxl import load_workbook
from io import BytesIO
from sqlalchemy.orm import Session
from app.models.record import Record
import os

TEMPLATE_PATH = r"e:\project\mz_mfp\3、中医门（急）诊诊疗信息页数据上传模板.xlsx"

class ExportService:
    def generate_excel(self, db: Session, record_id: int) -> BytesIO:
        record = db.query(Record).filter(Record.id == record_id).first()
        if not record:
            raise ValueError("Record not found")

        # Load Template
        if not os.path.exists(TEMPLATE_PATH):
             raise FileNotFoundError(f"Template not found at {TEMPLATE_PATH}")
             
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        # Map Headers to Column Index (0-based)
        # Row 1 is headers
        headers = {}
        for idx, cell in enumerate(ws[1]):
            if cell.value:
                headers[cell.value] = idx + 1 # 1-based col index for cell access or just use idx for list access

        # We will write to Row 2 (overwriting or appending? Template usually empty row 2)
        # Let's assume row 2 is for data.
        data_row = 2

        # Helper to set value by header name
        def set_val(header, value):
            if header in headers:
                col_idx = headers[header]
                ws.cell(row=data_row, column=col_idx, value=value)

        # 1. Base Info
        if record.base_info:
            bi = record.base_info
            set_val('XM', bi.xm)
            set_val('XB', bi.xb)
            set_val('CSRQ', bi.csrq)
            set_val('ZJHM', bi.zjhm)
            # ... add more mapping

        # 2. Diagnoses (Flatten)
        # MZZD_JBBM1..10 (TCM Disease Name), MZZD_ZYZD (Main TCM Name - usually first?)
        # Logic: MZZD_ZYZD is usually a specific column or the first one. 
        # Let's check the template analysis: 'MZZD_ZYZD', 'MZZD_JBBM', 'MZZD_JBBM1'...'10'
        # Assume tcm_diagnoses map to JBBM1..10
        tcm_diags = [d for d in record.diagnoses if d.diag_type == '1']
        tcm_diags.sort(key=lambda x: x.seq_no)
        
        for i, d in enumerate(tcm_diags):
            if i < 10:
                suffix = str(i + 1)
                set_val(f'MZZD_JBBM{suffix}', d.disease_name)
                # If there's code column
                # set_val(f'MZZD_JBDM{suffix}', d.disease_code) 

        # WM Diagnoses
        wm_diags = [d for d in record.diagnoses if d.diag_type == '2']
        wm_diags.sort(key=lambda x: x.seq_no)
         
        for i, d in enumerate(wm_diags):
            if i < 10:
                suffix = str(i + 1)
                set_val(f'MZZD_QTZD{suffix}', d.disease_name)

        # 3. Operations
        # SSCZMC1..5
        # Assuming operations map here
        # Note: Model has TcmOperation and Surgery. Need to decide which maps to SSCZ
        # Let's assume Surgery maps to SSCZ for now
        surgeries = sorted(record.surgeries, key=lambda x: x.seq_no)
        for i, op in enumerate(surgeries):
            if i < 5:
                suffix = str(i + 1)
                set_val(f'SSCZMC{suffix}', op.op_name)
                set_val(f'SSCZBM{suffix}', op.op_code)
                set_val(f'SSCZRQ{suffix}', op.op_date)

        # 4. Fee
        if record.fee_summary:
            f = record.fee_summary
            set_val('ZFY', f.zfy)
            set_val('ZFJE', f.zfje)
            set_val('YLFWF', f.ylfwf)
            # ... map all 42 items if needed

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

export_service = ExportService()
