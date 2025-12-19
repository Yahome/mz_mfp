from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterator, Optional

import openpyxl
from sqlalchemy import create_engine, text

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

from app.core.config import get_settings


RC_CODE_RE = re.compile(r"(RC\\d{3})")


@dataclass(frozen=True)
class DictRow:
    code: str
    name: str
    extra_code: Optional[str] = None
    merged_code: Optional[str] = None


def _as_str(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text_value = str(value).strip()
    return text_value if text_value else None


def _iter_two_col_sheet(ws) -> Iterator[DictRow]:
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = _as_str(row[0])
        name = _as_str(row[1]) if len(row) > 1 else None
        if not code or not name:
            continue
        yield DictRow(code=code, name=name)


def _iter_icd_sheet(ws) -> Iterator[DictRow]:
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = _as_str(row[0])
        extra = _as_str(row[1]) if len(row) > 1 else None
        merged = _as_str(row[2]) if len(row) > 2 else None
        name = _as_str(row[3]) if len(row) > 3 else None
        if not code or not name:
            continue
        yield DictRow(code=code, name=name, extra_code=extra, merged_code=merged)


def _get_set_name(ws, fallback: str) -> str:
    title = _as_str(ws.cell(1, 1).value)
    if not title:
        return fallback
    title = RC_CODE_RE.sub("", title).strip()
    return title or fallback


def _get_rc_set_code(ws, fallback: str) -> str:
    title = _as_str(ws.cell(1, 1).value) or ""
    match = RC_CODE_RE.search(title)
    if match:
        return match.group(1)
    return fallback


def _iter_sheet_dicts(wb) -> Iterator[tuple[str, str, Iterator[DictRow]]]:
    for sheet_name in wb.sheetnames:
        if sheet_name == "中医门（急）诊诊疗信息页数据接口标准":
            continue

        ws = wb[sheet_name]
        set_code: Optional[str] = None
        iterator: Optional[Iterator[DictRow]] = None

        if sheet_name.startswith("RC"):
            set_code = _get_rc_set_code(ws, fallback=sheet_name.split("-", 1)[0].strip())
            iterator = _iter_two_col_sheet(ws)
        elif "ICD-10" in sheet_name or "ICD10" in sheet_name:
            set_code = "ICD10"
            iterator = _iter_icd_sheet(ws)
        elif "ICD-9" in sheet_name or "手术操作编码" in sheet_name or "ICD-9-CM3" in sheet_name:
            set_code = "ICD9CM3"
            iterator = _iter_icd_sheet(ws)
        elif sheet_name.startswith("中医疾病"):
            set_code = "TCM_DISEASE"
            iterator = _iter_two_col_sheet(ws)
        elif sheet_name.startswith("中医证候"):
            set_code = "TCM_SYNDROME"
            iterator = _iter_two_col_sheet(ws)
        elif "肿瘤形态学" in sheet_name:
            set_code = "TUMOR_MORPHOLOGY"
            iterator = _iter_two_col_sheet(ws)
        elif sheet_name == "中草药类型":
            set_code = "HERB_TYPE"
            iterator = _iter_two_col_sheet(ws)
        elif sheet_name == "用药途径":
            set_code = "DRUG_ROUTE"
            iterator = _iter_two_col_sheet(ws)
        elif sheet_name == "国籍":
            set_code = "COUNTRY"
            iterator = _iter_two_col_sheet(ws)
        else:
            continue

        if not set_code or iterator is None:
            continue

        set_name = _get_set_name(ws, fallback=sheet_name)
        yield set_code, set_name, iterator


def main() -> None:
    parser = argparse.ArgumentParser(description="直接导入标准字典到 MySQL（会覆盖同 set_code 数据）")
    parser.add_argument(
        "--xlsx",
        default="2、中医门（急）诊诊疗信息页数据项采集接口标准.xlsx",
        help="标准 Excel 路径（相对于项目根或绝对路径）",
    )
    parser.add_argument("--batch-size", type=int, default=5000, help="每批写入行数（建议 2000-8000）")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"文件不存在: {xlsx_path}")

    settings = get_settings()
    engine = create_engine(settings.mysql_dsn, pool_pre_ping=True)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    for set_code, set_name, iterator in _iter_sheet_dicts(wb):
        with engine.begin() as conn:
            conn.execute(
                text(
                    "INSERT INTO dict_set(set_code,set_name,version,source) "
                    "VALUES(:set_code,:set_name,NULL,NULL) "
                    "ON DUPLICATE KEY UPDATE set_name=VALUES(set_name), version=VALUES(version), source=VALUES(source)"
                ),
                {"set_code": set_code, "set_name": set_name},
            )
            conn.execute(text("DELETE FROM dict_item WHERE set_code=:set_code"), {"set_code": set_code})

            seen_codes: set[str] = set()
            batch: list[dict] = []
            total = 0
            for row in iterator:
                if row.code in seen_codes:
                    continue
                seen_codes.add(row.code)
                batch.append(
                    {
                        "set_code": set_code,
                        "code": row.code,
                        "name": row.name,
                        "extra_code": row.extra_code,
                        "merged_code": row.merged_code,
                    }
                )
                if len(batch) >= args.batch_size:
                    conn.execute(
                        text(
                            "INSERT INTO dict_item(set_code,code,name,extra_code,merged_code,status,sort_no) "
                            "VALUES(:set_code,:code,:name,:extra_code,:merged_code,1,0)"
                        ),
                        batch,
                    )
                    total += len(batch)
                    batch = []
            if batch:
                conn.execute(
                    text(
                        "INSERT INTO dict_item(set_code,code,name,extra_code,merged_code,status,sort_no) "
                        "VALUES(:set_code,:code,:name,:extra_code,:merged_code,1,0)"
                    ),
                    batch,
                )
                total += len(batch)

        print(f"{set_code} {set_name}: {total} 行")


if __name__ == "__main__":
    main()
