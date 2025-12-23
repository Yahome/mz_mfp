from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Iterator, Optional

import openpyxl


RC_CODE_RE = re.compile(r"(RC\\d{3})")


@dataclass(frozen=True)
class DictRow:
    code: str
    name: str
    item_type: Optional[str] = None
    select_optional: Optional[str] = None


def _as_str(value: object) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text if text else None


def _sql_escape(value: str) -> str:
    return value.replace("\\\\", "\\\\\\\\").replace("'", "''")


def _iter_rc_sheet(ws) -> Iterator[DictRow]:
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = _as_str(row[0])
        name = _as_str(row[1]) if len(row) > 1 else None
        if not code or not name:
            continue
        yield DictRow(code=code, name=name)


def _iter_two_col_sheet(ws) -> Iterator[DictRow]:
    for row in ws.iter_rows(min_row=3, values_only=True):
        code = _as_str(row[0])
        name = _as_str(row[1]) if len(row) > 1 else None
        if not code or not name:
            continue
        yield DictRow(code=code, name=name)


def _iter_icd_sheet(ws) -> Iterator[DictRow]:
    for row in ws.iter_rows(min_row=3, values_only=True):
        main_code = _as_str(row[0])
        extra_code = _as_str(row[1]) if len(row) > 1 else None
        merged_code = _as_str(row[2]) if len(row) > 2 else None
        name = _as_str(row[3]) if len(row) > 3 else None
        item_type = _as_str(row[4]) if len(row) > 4 else None
        select_optional = _as_str(row[5]) if len(row) > 5 else None

        code = merged_code or main_code or extra_code
        if not code or not name:
            continue
        yield DictRow(code=code, name=name, item_type=item_type, select_optional=select_optional)


def _get_rc_set_code(ws) -> Optional[str]:
    title = _as_str(ws.cell(1, 1).value) or ""
    match = RC_CODE_RE.search(title)
    return match.group(1) if match else None


def _get_set_name(ws, fallback: str) -> str:
    title = _as_str(ws.cell(1, 1).value)
    if not title:
        return fallback
    title = title.strip()
    title = RC_CODE_RE.sub("", title).strip()
    return title or fallback


def _chunked(items: Iterable[DictRow], batch_size: int) -> Iterator[list[DictRow]]:
    batch: list[DictRow] = []
    for item in items:
        batch.append(item)
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch


def _write_sql(path: Path, sql: str) -> None:
    path.write_text(sql.strip() + "\n", encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="从标准 Excel 生成 dict_set/dict_item 导入 SQL（分批文件）")
    parser.add_argument(
        "--xlsx",
        default="2、中医门（急）诊诊疗信息页数据项采集接口标准.xlsx",
        help="标准 Excel 路径",
    )
    parser.add_argument("--outdir", default="backend/tmp/dict_sql", help="输出目录（会自动创建）")
    parser.add_argument("--batch-size", type=int, default=1000, help="每批 insert 行数")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    for sheet_name in wb.sheetnames:
        if sheet_name == "中医门（急）诊诊疗信息页数据接口标准":
            continue

        ws = wb[sheet_name]

        set_code: Optional[str] = None
        iterator: Optional[Iterator[DictRow]] = None

        if sheet_name.startswith("RC"):
            set_code = _get_rc_set_code(ws) or sheet_name.split("-", 1)[0].strip()
            iterator = _iter_rc_sheet(ws)
        elif "ICD-10" in sheet_name or "ICD10" in sheet_name:
            set_code = "ICD10"
            iterator = _iter_icd_sheet(ws)
        elif "ICD-9" in sheet_name or "手术操作编码" in sheet_name or "ICD-9-CM3" in sheet_name:
            set_code = "ICD9CM3"
            iterator = _iter_icd_sheet(ws)
        elif sheet_name.startswith("中医疾病"):
            set_code = "TCM_DISEASE"
            iterator = _iter_two_col_sheet(ws)
        elif sheet_name.startswith("中医证候"):
            set_code = "TCM_SYNDROME"
            iterator = _iter_two_col_sheet(ws)
        elif "肿瘤形态学" in sheet_name:
            set_code = "TUMOR_MORPHOLOGY"
            iterator = _iter_two_col_sheet(ws)
        elif sheet_name == "中草药类型":
            set_code = "HERB_TYPE"
            iterator = _iter_two_col_sheet(ws)
        elif sheet_name == "用药途径":
            set_code = "DRUG_ROUTE"
            iterator = _iter_two_col_sheet(ws)
        elif sheet_name == "国籍":
            set_code = "COUNTRY"
            iterator = _iter_two_col_sheet(ws)
        else:
            # 未识别的 sheet 暂不处理
            continue

        if not set_code or iterator is None:
            continue

        set_name = _get_set_name(ws, fallback=sheet_name)

        upsert_set_sql = (
            "INSERT INTO dict_set(set_code,set_name,version,source) "
            f"VALUES('{_sql_escape(set_code)}','{_sql_escape(set_name)}',NULL,NULL) "
            "ON DUPLICATE KEY UPDATE set_name=VALUES(set_name), version=VALUES(version), source=VALUES(source);"
        )
        _write_sql(outdir / f"{set_code}__00_upsert_set.sql", upsert_set_sql)

        delete_items_sql = f"DELETE FROM dict_item WHERE set_code='{_sql_escape(set_code)}';"
        _write_sql(outdir / f"{set_code}__01_delete_items.sql", delete_items_sql)

        for idx, batch in enumerate(_chunked(iterator, args.batch_size), start=1):
            values_sql = []
            for item in batch:
                item_type = "NULL" if item.item_type is None else f"'{_sql_escape(item.item_type)}'"
                select_optional = "NULL" if item.select_optional is None else f"'{_sql_escape(item.select_optional)}'"
                values_sql.append(
                    f"('{_sql_escape(set_code)}','{_sql_escape(item.code)}','{_sql_escape(item.name)}',{item_type},{select_optional},1,0)"
                )
            insert_sql = (
                "INSERT INTO dict_item(set_code,code,name,item_type,select_optional,status,sort_no) VALUES\\n"
                + ",\\n".join(values_sql)
                + "\\nON DUPLICATE KEY UPDATE name=VALUES(name), item_type=VALUES(item_type), select_optional=VALUES(select_optional), status=VALUES(status), sort_no=VALUES(sort_no);"
            )
            _write_sql(outdir / f"{set_code}__02_items_{idx:04d}.sql", insert_sql)

    print(f"已生成 SQL 分批文件到：{outdir}")


if __name__ == "__main__":
    main()
